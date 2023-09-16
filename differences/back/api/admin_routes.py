import os
from flask import jsonify, request, make_response

from flask_restful import Resource, abort
from flask_jwt_extended import create_access_token,get_jwt,jwt_required


#---import models
from models.users import User

#--- import db
from extensions import db


from flask import  render_template

from extensions import ph

#--- for comparing img 
import cv2
import numpy as np
import requests
from io import BytesIO

#--- for the excel
import openpyxl
from fileinput import filename
import requests
import os


class Signupator(Resource): #to get all questions, or create a new one
    method_decorators=[jwt_required()]
    def post(self):
        print("Asd")
        data = request.json
        user = data.get("user")
        password = data.get("password")
        rep_password = data.get("rep_password")
       
        if not user or not password:
            return make_response(jsonify({"msg" :  "missing data"}),400)
        if password != rep_password:
            return make_response(jsonify({"msg" : "passwords dont match"}),400)

        existe = User.query.filter_by(user=user).first()
    
        if existe: 
            return make_response(jsonify({"msg" : "User exists"}),400)

        hash = ph.hash(password)
        
        addUsuario = User(user = user, password=hash)
        print(addUsuario.serialize())
        db.session.add(addUsuario)
        db.session.commit()

        return jsonify({"msg" : "user registered"})
    

class Loginator(Resource): #to get all questions, or create a new one
    def post(self):
            data = request.json
            user = data.get("user")
            password = data.get("password")
            print(os.environ.get("JWT_SECRET_KEY"))
            
            if not user or not password:
                return make_response(jsonify({"msg" :  "missing data"}),400)
            
            hash = ph.hash(password)
            user = User.query.filter_by(user=user).first()

            if not user:
                return make_response(jsonify({"msg": "Wrong username or password"}), 400)
            else:
                try:
                    if(not ph.verify(user.password,password)):
                       
                        return make_response(jsonify({"msg": "Wrong username or password"}), 400)
                except: 
                    
                    return make_response(jsonify({"msg": "Wrong username or password"}), 400)
            
            token = create_access_token(identity=user.id)
            return jsonify({ "token": token, "msg":"login success"})
    
class Token_validator(Resource):
    method_decorators=[jwt_required()]
    def get(self):
        pass




#Examples of calls


class Compareimg(Resource): 
    
    def get(self):
        pass

    def post(self):
        
        data = request.json
        url1 = data.get("url1") # original img
        url2 = data.get("url2") # img with differences marked
        url3 = data.get("url3") # img with differences not marked
        
        try:
            response1 = requests.get(url1)
            response2 = requests.get(url2)
            response3 = requests.get(url3)

            image_data1 = BytesIO(response1.content)
            image_data2 = BytesIO(response2.content)
            image_data3 = BytesIO(response3.content)

            image_array1 = np.asarray(bytearray(image_data1.read()), dtype=np.uint8)
            image_array2 = np.asarray(bytearray(image_data2.read()), dtype=np.uint8)
            image_array3 = np.asarray(bytearray(image_data3.read()), dtype=np.uint8)

            image1 = cv2.imdecode(image_array1, cv2.IMREAD_COLOR)
            image2 = cv2.imdecode(image_array2, cv2.IMREAD_COLOR)
            image3 = cv2.imdecode(image_array3, cv2.IMREAD_COLOR)
        except cv2.error:
            return(make_response(jsonify({"msg" : "An error occured, check the img url"}),400))
                
        except:
            return(make_response(jsonify({"msg" : "An error occured while retrieving the images"}),400))

        
        difference = cv2.absdiff(image1, image2)
        gray = cv2.cvtColor(difference, cv2.COLOR_BGR2GRAY)
        _, thresholded = cv2.threshold(gray, 30, 255, cv2.THRESH_BINARY)
        
        contours, _ = cv2.findContours(thresholded, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
       

        cv2.drawContours(image3, contours, -1, (0, 255, 0), 2)
        #cv2.imshow('Image with Contours', image3)
        #cv2.imshow('og', image1)
        #cv2.imshow('differences', image2)

        cv2.waitKey(0)
        cv2.destroyAllWindows()

        center_coordinates = []


        for contour in contours:
            # Calculate the moments of the contour
            M = cv2.moments(contour)
            
            # Calculate the centroid (center) of the contour
            if M["m00"] != 0:
                cX = int(M["m10"] / M["m00"])
                cY = int(M["m01"] / M["m00"])
            else:
                # Handle the case where the contour has no area (division by zero)
                cX, cY = 0, 0
            
            # Append the center coordinates to the list
            if cX != 0 and cY != 0:
                center_coordinates.append((cX, cY))

            # Print the center coordinates of each contour
        for i, (cX, cY) in enumerate(center_coordinates):
            print(f"Contour {i + 1}: Center = ({cX}, {cY})")

       
        return jsonify({"coordinates" : center_coordinates})
    

class Img_upload(Resource):
    method_decorators=[jwt_required()]
    def post(self):
        token_dic = get_jwt()
        token = token_dic.get("jti")
        
        file = request.files['file']
        file.save(file.filename)
        wb = openpyxl.load_workbook(file.filename)
        ws = wb.active

        has_text = "a"
        row_ammount = 0
        c = 2

        while (has_text):
            has_text = ws.cell(row=c, column=1).value
            
            if(has_text):
                row_ammount += 1
            c += 1

        if(row_ammount) < 1:
            return make_response(jsonify({"msg" : "Error, input at least one row of urls"}),400)
        
        for i in range(2,row_ammount + 2):
        
            difference_to_add = {
                    "url1": "",
                    "url2": "",
                    "url3" : ""
                }
    
            difference_to_add["url1"] = ws.cell(row=i,column=1).value    
            difference_to_add["url2"] = ws.cell(row=i,column=2).value        
            difference_to_add["url3"] = ws.cell(row=i,column=3).value 
                
            difference_to_add["date"] = ws.cell(row=i,column=4).value.strftime('%d/%m/%Y')
            
            headers = {'Authorization': 'Bearer {}'.format(token)}
            
            response = requests.post(os.environ.get("VITE_BACKEND_URL", False) + "auth/compare", json=difference_to_add, headers=headers)

            print("asd")
            print(response.json())


        return jsonify({"msg": "added"})
